#!/usr/bin/env python3
"""
Extract data from comptes.xlsx into JSON files for import into the database.
Only uses the "Comptes XXXX" sheets (2019-2026).

Usage: python3 prisma/extract-excel.py

Output files in prisma/data/:
  - categories.json
  - transactions-bnp.json
"""

import json
import os
import datetime
import openpyxl

EXCEL_PATH = os.path.expanduser("~/Downloads/comptes.xlsx")
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "data")

PALETTE = [
    "#6366f1", "#22c55e", "#f59e0b", "#ef4444", "#8b5cf6",
    "#ec4899", "#06b6d4", "#14b8a6", "#64748b", "#10b981",
    "#059669", "#0ea5e9", "#f43f5e", "#78716c", "#a855f7",
    "#3b82f6", "#d946ef", "#84cc16",
]

MONTH_MAP = {
    "JANVIER": 1, "FÉVRIER": 2, "MARS": 3, "AVRIL": 4,
    "MAI": 5, "JUIN": 6, "JUILLET": 7, "AOÛT": 8, "AOUT": 8,
    "SEPTEMBRE": 9, "OCTOBRE": 10, "NOVEMBRE": 11, "DÉCEMBRE": 12,
    "DECEMBRE": 12, "DÉCEMBRE 2K18": 12,
    "Janvier": 1, "Février": 2, "Mars": 3, "Avril": 4,
    "Mai": 5, "Juin": 6, "Juillet": 7, "Août": 8,
    "Septembre": 9, "Octobre": 10, "Novembre": 11, "Décembre": 12,
}


def is_numeric(val):
    if val is None:
        return False
    if isinstance(val, (int, float)):
        return True
    if isinstance(val, str):
        try:
            float(val.replace(",", "."))
            return True
        except ValueError:
            return False
    return False


def to_float(val):
    if isinstance(val, (int, float)):
        return round(float(val), 2)
    if isinstance(val, str):
        return round(float(val.replace(",", ".")), 2)
    return 0.0


def parse_date(val, year, month):
    if val is None or val == "" or val == "Auto":
        return None
    if isinstance(val, datetime.datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, datetime.date):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, datetime.time):
        return None
    if isinstance(val, str):
        s = val.strip()
        if s.lower() == "auto" or s == "":
            return None
        if s.lower().startswith("le "):
            try:
                day = int(s.split()[1])
                return f"{year}-{month:02d}-{day:02d}"
            except (ValueError, IndexError):
                return None
    return None


def parse_status(val):
    if val is None:
        return "PENDING"
    s = str(val).strip()
    if s == "Oui":
        return "COMPLETED"
    if s == "Non":
        return "PENDING"
    if s.upper() in ("ANNULÉ", "ANNULE"):
        return "CANCELLED"
    return "PENDING"


def is_skip_row(amount_val, label_val):
    if not is_numeric(amount_val):
        return True
    if label_val is not None:
        label_str = str(label_val).strip()
        if "Somme initiale" in label_str:
            return True
        if label_str.lower() in ("", "remboursé"):
            return True
    return False


def detect_month_columns(ws):
    """Detect month positions from row 1 and determine cols_per_month for each."""
    raw = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val is not None and val in MONTH_MAP:
            raw.append((col, MONTH_MAP[val]))

    # Compute cols_per_month for each month using gap to next month
    result = []
    for i, (col, month_num) in enumerate(raw):
        if i + 1 < len(raw):
            gap = raw[i + 1][0] - col
        else:
            # Last month: infer from headers
            gap = _count_month_cols_from_headers(ws, col)
        result.append((col, month_num, gap))
    return result


def _count_month_cols_from_headers(ws, start_col):
    """Count how many columns this month block has by reading row 2 headers."""
    count = 0
    for c in range(start_col, start_col + 8):
        val = ws.cell(row=2, column=c).value
        if val is not None:
            count += 1
        elif count > 0:
            break
    return max(count, 3)


def _get_month_layout(ws, start_col, cols):
    """Determine layout type from row 2 headers for a specific month block."""
    headers = []
    for c in range(start_col, start_col + cols):
        val = ws.cell(row=2, column=c).value
        h = str(val).strip().replace("\n", " ") if val else ""
        headers.append(h.lower())

    has_cat = any("catégorie" in h or "categorie" in h for h in headers)
    has_subcat = any("sous" in h for h in headers)
    has_date_compte = any("date" in h and "compte" in h for h in headers)

    if cols == 3:
        return "simple"          # Amount, Label, Status
    if has_subcat:
        return "with_subcat"     # Amount, Label, Date, Category, SubCategory, Status
    if has_cat and not has_date_compte:
        return "with_cat"        # Amount, Label, Date, Category, Status
    if has_date_compte and has_cat:
        return "date_compte_cat" # Amount, Label, Date, Category, Status (2023-2024 have "Total sur compte" not "Date compte")
    if has_date_compte:
        return "date_compte"     # Amount, Label, Date, DateCompte, Status
    return "simple"


def extract_transactions_sheet(ws, sheet_name):
    """Extract transactions from a Comptes XXXX sheet."""
    transactions = []
    month_columns = detect_month_columns(ws)
    if not month_columns:
        print(f"  WARNING: No months found in {sheet_name}")
        return transactions

    base_year = int(sheet_name.split()[-1])

    print(f"  {sheet_name}: {len(month_columns)} months detected")

    for pos_idx, (start_col, month_num, cols) in enumerate(month_columns):
        # Determine actual year
        if base_year == 2019 and pos_idx < 3:
            year = 2018
        else:
            year = base_year

        layout = _get_month_layout(ws, start_col, cols)

        for row in range(5, ws.max_row + 1):
            amount_val = ws.cell(row=row, column=start_col).value
            label_val = ws.cell(row=row, column=start_col + 1).value

            if is_skip_row(amount_val, label_val):
                continue

            amount = to_float(amount_val)
            label = str(label_val).strip() if label_val else ""
            if not label:
                continue

            date_val = None
            category = None
            subcategory = None
            status_val = None

            if layout == "simple":
                # Amount, Label, Status
                status_val = ws.cell(row=row, column=start_col + 2).value
            elif layout == "date_compte":
                # Amount, Label, Date, DateCompte, Status
                date_val = ws.cell(row=row, column=start_col + 2).value
                status_val = ws.cell(row=row, column=start_col + 4).value
            elif layout == "with_cat" or layout == "date_compte_cat":
                # Amount, Label, Date, Category, Status
                date_val = ws.cell(row=row, column=start_col + 2).value
                category = ws.cell(row=row, column=start_col + 3).value
                status_val = ws.cell(row=row, column=start_col + 4).value
            elif layout == "with_subcat":
                # Amount, Label, Date, Category, SubCategory, Status
                date_val = ws.cell(row=row, column=start_col + 2).value
                category = ws.cell(row=row, column=start_col + 3).value
                subcategory = ws.cell(row=row, column=start_col + 4).value
                status_val = ws.cell(row=row, column=start_col + 5).value

            parsed_date = parse_date(date_val, year, month_num)
            status = parse_status(status_val)

            cat_str = str(category).strip() if category else None
            if cat_str in (None, "", "None", "Reste", "-", "Catégorie"):
                cat_str = None
            # Normalize accented variants
            if cat_str == "Économies":
                cat_str = "Economies"
            subcat_str = str(subcategory).strip() if subcategory else None
            if subcat_str in (None, "", "None", "Reste", "-"):
                subcat_str = None

            tx = {
                "year": year,
                "month": month_num,
                "amount": amount,
                "label": label,
                "date": parsed_date,
                "status": status,
            }
            if cat_str:
                tx["category"] = cat_str
            if subcat_str:
                tx["subcategory"] = subcat_str

            transactions.append(tx)

    return transactions


def build_categories_from_transactions(transactions):
    """Build categories list from category/subcategory values found in transactions."""
    cat_subs = {}  # { category_name: set(subcategory_names) }

    for tx in transactions:
        cat = tx.get("category")
        if not cat:
            continue
        if cat not in cat_subs:
            cat_subs[cat] = set()
        sub = tx.get("subcategory")
        if sub:
            cat_subs[cat].add(sub)

    # Sort categories alphabetically, build output
    categories = []
    for name in sorted(cat_subs.keys()):
        color = PALETTE[len(categories) % len(PALETTE)]
        categories.append({
            "name": name,
            "subcategories": sorted(cat_subs[name]),
            "color": color,
        })

    # Add "Non catégorisé" for transactions without category
    categories.append({
        "name": "Non catégorisé",
        "subcategories": [],
        "color": "#9ca3af",
    })

    return categories


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print(f"Loading {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # Extract transactions from Comptes 2019-2026 only
    print("\n--- Extracting BNP transactions ---")
    all_transactions = []
    for year in range(2019, 2027):
        sheet_name = f"Comptes {year}"
        if sheet_name in wb.sheetnames:
            txs = extract_transactions_sheet(wb[sheet_name], sheet_name)
            all_transactions.extend(txs)
            print(f"    {sheet_name}: {len(txs)} transactions")

    # Build categories from transaction data
    print("\n--- Building categories from transaction data ---")
    categories = build_categories_from_transactions(all_transactions)
    print(f"  Found {len(categories)} categories:")
    for cat in categories:
        subs = f" → {cat['subcategories']}" if cat['subcategories'] else ""
        print(f"    {cat['name']}{subs}")

    # Write JSON files
    def write_json(filename, data):
        path = os.path.join(OUTPUT_DIR, filename)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"  Written {path} ({len(data)} entries)")

    print("\n--- Writing JSON files ---")
    write_json("categories.json", categories)
    write_json("transactions-bnp.json", all_transactions)

    # Summary
    categorized = sum(1 for t in all_transactions if "category" in t)
    uncategorized = len(all_transactions) - categorized
    print(f"\n=== SUMMARY ===")
    print(f"  Categories: {len(categories)}")
    print(f"  Total transactions: {len(all_transactions)}")
    print(f"  With category: {categorized}")
    print(f"  Without category (to be categorized by Claude): {uncategorized}")

    # Breakdown by year
    from collections import Counter
    year_counts = Counter(t["year"] for t in all_transactions)
    for y in sorted(year_counts):
        cat_count = sum(1 for t in all_transactions if t["year"] == y and "category" in t)
        print(f"    {y}: {year_counts[y]} tx ({cat_count} categorized)")


if __name__ == "__main__":
    main()
